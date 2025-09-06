from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from flask import Flask, render_template, request, send_from_directory
from fpdf import FPDF
import csv
from datetime import datetime
import random
import os

app = Flask(__name__)
app.config['RECEIPT_FOLDER'] = os.path.join('static', 'receipts')  # ✅ correct for your structure


# ---------------------------
# Assign agent randomly
# ---------------------------
def assign_agent():
    agents = [
        {"Name": "Ravi Sharma", "Contact": "9876543210"},
        {"Name": "Amit Verma", "Contact": "9876543211"},
        {"Name": "Mohit Kumar", "Contact": "9876543212"},
        {"Name": "Rohit Reddy", "Contact": "9876543213"},
        {"Name": "Raj Patel", "Contact": "9876543214"},
        {"Name": "Amit Joshi", "Contact": "9876543215"},
        {"Name": "Karan Gupta", "Contact": "9876543216"},
        {"Name": "Alok Mishra", "Contact": "9876543217"},
        {"Name": "Suresh Rana", "Contact": "9876543218"},
        {"Name": "Deepak Jain", "Contact": "9876543219"},
        {"Name": "Arjun Malhotra", "Contact": "9876543220"},
        {"Name": "Rohit Sinha", "Contact": "9876543221"},
        {"Name": "Yash Thakur", "Contact": "9876543222"},
        {"Name": "Nikhil Agarwal", "Contact": "9876543223"},
        {"Name": "Vivek Deshmukh", "Contact": "9876543224"},
        {"Name": "Tarun Kapoor", "Contact": "9876543225"},
        {"Name": "Harsh Tiwari", "Contact": "9876543226"},
        {"Name": "Siddharth Mehta", "Contact": "9876543227"},
        {"Name": "Ankit Chauhan", "Contact": "9876543228"},
        {"Name": "Shubham Tripathi", "Contact": "9876543229"},
        {"Name": "Abhishek Rawat", "Contact": "9876543230"},
        {"Name": "Lakshya Nanda", "Contact": "9876543231"},
        {"Name": "Vikas Kulkarni", "Contact": "9876543232"},
        {"Name": "Gaurav Bhatt", "Contact": "9876543233"},
        {"Name": "Nitin Joshi", "Contact": "9876543234"},
        {"Name": "Manish Srivastava", "Contact": "9876543235"},
        {"Name": "Aditya Chauhan", "Contact": "9876543236"},
        {"Name": "Rahul Saxena", "Contact": "9876543237"},
        {"Name": "Aayush Bansal", "Contact": "9876543238"},
        {"Name": "Mayank Pandey", "Contact": "9876543239"},
    ]
    return random.choice(agents)

# ---------------------------
# Save service request to Excel
# ---------------------------
def save_user_request(name, address, contact, email, appliance, problem, agent):
    filename = 'user_requests.xlsx'
    file_exists = os.path.isfile(filename)
    serial_no = 1

    if not file_exists:
        wb = Workbook()
        ws = wb.active
        ws.append(["S.No.", "Timestamp", "Name", "Address", "Contact", "Email", "Appliance", "Problem", "Agent"])
    else:
        wb = load_workbook(filename)
        ws = wb.active
        serial_no = ws.max_row  # since header is row 1

    ws.append([serial_no, datetime.now(), name, address, contact, email, appliance, problem, agent["Name"]])

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)

# ---------------------------
# Generate PDF Receipt
# ---------------------------
def generate_pdf_receipt(name, address, contact, email, appliance, problem, agent):
    if not os.path.exists(app.config['RECEIPT_FOLDER']):
        os.makedirs(app.config['RECEIPT_FOLDER'])

    pdf = FPDF()
    pdf.add_page()

    # Header - Company Name
    pdf.set_font("Arial", 'B', 20)
    pdf.set_text_color(25, 25, 112)  # Navy Blue
    pdf.cell(200, 10, txt="AppliFix", ln=True, align='C')

    # Tagline
    pdf.set_font("Arial", '', 12)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(200, 10, txt="Your Trusted Home Appliance Service Partner", ln=True, align='C')

    pdf.ln(10)

    # Receipt Title
    pdf.set_font("Arial", 'B', 14)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(200, 10, txt="SERVICE BOOKING RECEIPT", ln=True, align='C')
    pdf.ln(5)

    # Horizontal line
    pdf.set_draw_color(100, 100, 100)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(8)

    # Body
    def field(label, value):
        pdf.set_font("Arial", 'B', 12)
        pdf.set_text_color(50, 50, 50)
        pdf.cell(55, 10, f"{label}:", ln=0)
        pdf.set_font("Arial", '', 12)
        pdf.set_text_color(30, 30, 30)
        pdf.multi_cell(0, 10, value)

    field("Customer Name", name)
    field("Address", address)
    field("Contact Number", contact)
    field("Email", email)
    field("Appliance", appliance)
    field("Problem Description", problem)
    field("Assigned Agent", f"{agent['Name']} ({agent['Contact']})")

    pdf.ln(5)

    # Timestamp
    pdf.set_font("Arial", 'I', 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 10, txt=f"Booking Time: {datetime.now().strftime('%d-%m-%Y  %I:%M %p')}", ln=True)

    # Footer
    pdf.set_y(-25)
    pdf.set_font("Arial", 'I', 9)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 10, txt="Thank you for trusting MySmartDwell. We bring care to your repair.", align='C')

    # Save
    filename = f"receipt_{name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    pdf_path = os.path.join(app.config['RECEIPT_FOLDER'], filename)
    pdf.output(pdf_path)

    return filename
# ---------------------------
# Feedback Excel Save
# ---------------------------
def collect_feedback(name, agent, rating, review):
    filename = 'feedback.xlsx'
    file_exists = os.path.isfile(filename)
    serial_no = 1

    if not file_exists:
        wb = Workbook()
        ws = wb.active
        ws.append(["S.No.", "Timestamp", "Customer Name", "Agent Name", "Rating", "Review"])
    else:
        wb = load_workbook(filename)
        ws = wb.active
        serial_no = ws.max_row

    ws.append([serial_no, datetime.now(), name, agent, rating, review])

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)

# ---------------------------
# Routes
# ---------------------------
@app.route('/')
def welcome():
    return render_template('index.html')

@app.route('/service')
def service():
    return render_template('service.html')

@app.route('/request', methods=['POST'])
def request_service():
    user_name = request.form['name']
    user_address = request.form['address']
    user_contact = request.form['contact']
    user_email = request.form['email']
    appliance_type = request.form['appliance']
    problem_desc = request.form['problem']

    agent = assign_agent()
    save_user_request(user_name, user_address, user_contact, user_email, appliance_type, problem_desc, agent)

    pdf_filename = generate_pdf_receipt(user_name, user_address, user_contact, user_email, appliance_type, problem_desc, agent)

    return render_template('confirmation.html',
                           name=user_name,
                           appliance=appliance_type,
                           problem=problem_desc,
                           agent=agent,
                           pdf_filename=pdf_filename)

@app.route('/download/<path:filename>')
def download_receipt(filename):
    return send_from_directory(app.config['RECEIPT_FOLDER'], filename, as_attachment=True)

@app.route('/feedback', methods=['GET', 'POST'])
def feedback():
    if request.method == 'POST':
        name = request.form['name']
        agent = request.form['agent']
        rating = request.form['rating']
        review = request.form['review']
        collect_feedback(name, agent, rating, review)
        return render_template('thank_you.html', name=name)

    name = request.args.get('name', '')
    agent = request.args.get('agent', '')
    return render_template('feedback.html', name=name, agent=agent)

@app.route('/all-feedbacks')
def all_feedbacks():
    feedback_list = []
    try:
        wb = load_workbook("feedback.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
            if len(row) >= 6:
                feedback_list.append({
                    "name": row[2],
                    "agent": row[3],
                    "rating": row[4],
                    "review": row[5]
                })
    except FileNotFoundError:
        pass  # no feedback file yet

    return render_template('feedbacks.html', feedbacks=feedback_list)

if __name__ == '__main__':
    app.run(debug=True, port=5001)

